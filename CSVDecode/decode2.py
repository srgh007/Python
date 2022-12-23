import csv
import fileinput
import locale
import os
import re
import sys
from os import walk

import pandas
from openpyxl import load_workbook

if not sys.warnoptions:
    import warnings

    warnings.simplefilter("ignore")
# import pyodbc

# cnxn = pyodbc.connect(
#     "Driver={SQL Server Native Client 11.0};"
#     "Server=server_name;"
#     "Database=db_name;"
#     "Trusted_Connection=yes;"
# )
# cursor = cnxn.cursor()
# cursor.execute("SELECT * FROM Table")

# for row in cursor:
#     print("row = %r" % (row,))

locale.setlocale(locale.LC_ALL, "ru_RU")

outfile = ""
for (dirpath, dirnames, filenames) in walk("E:\\Python\\CSVDecode\\IN"):
    print(filenames)
    for file in filenames:
        if "xls" in file:
            outfile = os.path.join(dirpath, file)
            print(outfile)

wb_form = load_workbook(filename=outfile)
wb_val = load_workbook(filename=outfile, data_only=True)

sheet_val = wb_val["Тело"]

cell_vals = sheet_val["{}4".format(chr(65))].value
print(cell_vals)

# listprms = []

listcols = []
dictpage = {}

# for row in records:
#     db[row["title"]] = dict(title=row["title"], id=row["id"])
# print()


def listOfparams():
    maxrow = 0
    for i in range(65, 90):
        listcols.append(sheet_val["{}4".format(chr(i))].value)
        j = 5  # for j in range(5, 100):
        for j in range(5, sheet_val.max_row):
            # while sheet_val["{}{}".format(chr(i), j)].value != None:
            # print(sheet_val["{}{}".format(chr(i), j)].value)
            if sheet_val["{}{}".format(chr(i), j)].value != None:
                dictpage[(sheet_val["{}4".format(chr(i))].value, j)] = sheet_val[
                    "{}{}".format(chr(i), j)
                ].value
            else:
                dictpage[(sheet_val["{}4".format(chr(i))].value, j)] = ""
    #             j += 1
    #             maxrow = j
    #         else:
    #             break
    # return maxrow
    # for g in range(65, 66):
    #     for i in range(65, 90):
    #         if sheet_val["{}{}4".format(chr(g), chr(i))].value != None:
    #             listprms.append(sheet_val["{}{}4".format(chr(g), chr(i))].value)
    #             # x = sheet_val["{}{}5".format(chr(g), chr(i))].value
    #             for j in range(5, 10):
    #                 print(sheet_val["{}{}{}".format(chr(g), chr(i), j)].value)
    #                 if sheet_val["{}{}{}".format(chr(g), chr(i), j)].value != None:
    #                     dictcolumns[
    #                         sheet_val["{}{}4".format(chr(g), chr(i))]
    #                         .value : sheet_val["{}{}{}".format(chr(g), chr(i), j)]
    #                         .value
    #                     ]
    #                 else:
    #                     continue
    #         else:
    #             return

    # for i in range(65, 90):
    #     if sheet_val["B{}4".format(chr(i))].value != None:
    #         listprms.append(sheet_val["A{}4".format(chr(i))].value)
    #     else:
    #         return listprms


listOfparams()

# print(xrow)

dcttple = tuple(dictpage.keys())

dictcolumns = [dictpage[("Серийный номер", n)] for n in range(5, sheet_val.max_row)]

print(dictcolumns)

dictrows = [dictpage[(z, 5)] for z in listcols]

print(dictrows)
# print(dictpage[("Операционный сервис", 5)])

# onecol = [dictpage[("Идентификационный номер (маркировка)"), x] for x in dcttple[-1]]
# print(onecol)
# print(dictpage)


# print(dcttple[0])
# print(list(dcttple[0,]))
# print(dictcolumns[("Серийный номер", 8)])

# print()

# type: ignore
csvlist = {}
maplist = {}
with open("CSVDecode\\mapping.csv", encoding="UTF8") as csvfile:
    reader = csv.DictReader(csvfile, delimiter=";")

    for row in reader:
        csvlist[row["fieldName"]] = row["Name"]
        maplist[row["integrName"]] = row["fieldName"]
    # print(csvlist["cModelName"])
    # print(maplist["Подтип КЭ"])

del maplist[""]

outfile = ""
for (dirpath, dirnames, filenames) in walk("E:\\Python\\CSVDecode\\OUT"):
    outfile = os.path.join(dirpath, filenames[0])

df = pandas.read_csv(outfile, sep=r";", index_col=False)

df.rename(columns=maplist, inplace=True)
df.to_csv(outfile, sep=";", index=False)
